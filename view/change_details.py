# import openpyxl and tkinter modules
from openpyxl import *
from tkinter import *

wb = load_workbook(r'C:\Users\User\Documents\שרי\Israel_It\view\task.xlsx')

# create the sheet object
sheet = wb.active


class Change_Details:
    def __init__(self):
        # create a GUI window
        self.root = Tk()
        self.n = 10000
        # set the background colour of GUI window
        self.root.configure(background='light green')

        # set the title of GUI window
        self.root.title("Change research details")

        # set the configuration of GUI window
        self.root.geometry("500x300")

        self.excel()

        # create a Form label
        heading = Label(self.root, text="Details", bg="light green")

        B = Label(self.root, text="Start Time", bg="green")

        C = Label(self.root, text="End Time", bg="green")

        # create a Contact No. label
        interval_no = Label(self.root, text="Interval No.", bg="light green")

        # create a Semester label
        interval = Label(self.root, text="Duration of the interval", bg="light green")

        # create a Form No. lable
        waiting = Label(self.root, text="waiting time", bg="light green")

        # grid method is used for placing
        # the widgets at respective positions
        # in table like structure .
        heading.grid(row=0, column=1)
        B.grid(row=2, column=0)
        C.grid(row=3, column=0)
        interval_no.grid(row=4, column=0)
        interval.grid(row=5, column=0)
        waiting.grid(row=6, column=0)

        # create a text entry box
        # for typing the information
        self.B_field = Entry(self.root)
        self.C_field = Entry(self.root)
        self.interval_no_field = Entry(self.root)
        self.interval_field = Entry(self.root)
        self.waiting_field = Entry(self.root)

        # bind method of widget is used for
        # the binding the function with the events

        # whenever the enter key is pressed
        # then call the focus1 function
        self.B_field.bind("<Return>", self.focus1)

        # whenever the enter key is pressed
        # then call the focus2 function
        self.C_field.bind("<Return>", self.focus2)

        # whenever the enter key is pressed
        # then call the focus3 function
        self.interval_no_field.bind("<Return>", self.focus3)

        # whenever the enter key is pressed
        # then call the focus4 function
        self.interval_field.bind("<Return>", self.focus4)

        # whenever the enter key is pressed
        # then call the focus5 function
        self.waiting_field.bind("<Return>", self.focus5)

        # grid method is used for placing
        # the widgets at respective positions
        # in table like structure .
        self.B_field.grid(row=1, column=1, ipadx="100")
        self.C_field.grid(row=2, column=1, ipadx="100")
        self.interval_no_field.grid(row=3, column=1, ipadx="100")
        self.interval_field.grid(row=4, column=1, ipadx="100")
        self.waiting_field.grid(row=5, column=1, ipadx="100")

        # call excel function
        self.excel()

        # create a Submit Button and place into the root window
        submit = Button(self.root, text="Submit", fg="Black",
                        bg="Red", command=self.insert)
        submit.grid(row=8, column=1)

        # start the GUI
        self.root.mainloop()

        # globally declare wb and sheet variable

        # opening the existing excel file

    def excel(self):
        # resize the width of columns in
        # excel spreadsheet

        sheet.column_dimensions['A'].width = 30
        sheet.column_dimensions['B'].width = 10
        sheet.column_dimensions['C'].width = 10
        sheet.column_dimensions['D'].width = 20
        sheet.column_dimensions['E'].width = 20
        sheet.column_dimensions['F'].width = 40
        sheet.column_dimensions['G'].width = 50

        # write given data to an excel spreadsheet
        # at particular location
        sheet.cell(row=1, column=2).value = "Duration of the interval"
        sheet.cell(row=1, column=3).value = "waiting time"
        for i in range(2, self.n + 2):
            sheet.cell(row=i, column=1).value = f'interval number {i - 2}'

    # Function to set focus (cursor)
    def focus1(self, event):
        # set focus on the course_field box
        self.course_field.focus_set()

    # Function to set focus
    def focus2(self, event):
        # set focus on the sem_field box
        self.sem_field.focus_set()

    # Function to set focus
    def focus3(self, event):
        # set focus on the form_no_field box
        self.form_no_field.focus_set()

    # Function to set focus
    def focus4(self, event):
        # set focus on the contact_no_field box
        self.contact_no_field.focus_set()

    # Function to set focus
    def focus5(self, event):
        # set focus on the email_id_field box
        self.email_id_field.focus_set()

    # Function for clearing the
    # contents of text entry boxes
    def clear(self):
        # clear the content of text entry box
        self.B_field.delete(0, END)
        self.C_field.delete(0, END)
        self.interval_no_field.delete(0, END)
        self.interval_field.delete(0, END)
        self.waiting_field.delete(0, END)

    # Function to take data from GUI
    # window and write to an excel file
    def insert(self):
        # if user not fill any entry
        # then print "empty input"
        if (self.interval_no_field.get() == "" and
                self.interval_field.get() == "" and
                self.waiting_field.get() == ""):

            print("empty input")

        else:

            # assigning the max row and max column
            # value upto which data is written
            # in an excel sheet to the variable
            current_row = sheet.max_row
            current_column = sheet.max_column

            # get method returns current text
            # as string which we write into
            # excel spreadsheet at particular location

            sheet.cell(row=current_row + 1, column=1).value = self.B_field.get()
            sheet.cell(row=current_row + 1, column=2).value = self.C_field.get()
            sheet.cell(row=current_row + 1, column=3).value = self.interval_no_field.get()
            sheet.cell(row=current_row + 1, column=4).value = self.interval_field.get()
            sheet.cell(row=current_row + 1, column=5).value = self.waiting_field.get()

            # save the file
            wb.save(r'C:\Users\User\Documents\שרי\Israel_It\view\task.xlsx')

            # set focus on the name_field box
            self.B_field.focus_set()

            # call the clear() function
            self.clear()
